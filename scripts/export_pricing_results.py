"""
Export pricing test results to Excel.

Usage:
  python scripts/export_pricing_results.py
  python scripts/export_pricing_results.py --tests scripts/pricing_test_cases.json --out outputs/pricing_audit.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# reuse the estimator from the test script
sys.path.insert(0, str(Path(__file__).parent))
from test_pricing_tool import estimate_fedex, estimate_ups  # noqa: E402

# ---------------------------------------------------------------------------
# Colors
# ---------------------------------------------------------------------------
NAVY        = "1B2B4B"
NAVY_LIGHT  = "2E4070"
GREEN_PASS  = "D6F0D6"
RED_FAIL    = "FAD4D4"
AMBER_ERR   = "FFF3CC"
WHITE       = "FFFFFF"
GRAY_ROW    = "F5F6FA"
HEADER_TXT  = "FFFFFF"
GREEN_TXT   = "1A6B1A"
RED_TXT     = "8B0000"
AMBER_TXT   = "7A5200"

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=None, size=10, name="Calibri"):
    return Font(bold=bold, color=color or "000000", size=size, name=name)

def _border():
    thin = Side(style="thin", color="D0D4E0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------
COLS = [
    # (header, width, align)
    ("TC#",                  6,  "center"),
    ("Origin",               8,  "center"),
    ("Dest",                 8,  "center"),
    ("Service",             16,  "left"),
    ("Program",             10,  "center"),
    ("Delivery",            11,  "center"),
    ("Wt (lb)",              8,  "center"),
    ("Dims (L×W×H)",        14,  "center"),
    ("Flags",               18,  "left"),
    ("Declared $",          10,  "center"),
    ("Markup %",             9,  "center"),
    # Tool outcome
    ("Tool Zone",            9,  "center"),
    ("Tool Bill Wt",        10,  "center"),
    ("Tool Total",          11,  "center"),
    # Python outcome
    ("Py Zone",              9,  "center"),
    ("Py Bill Wt",          10,  "center"),
    ("Py Published Rate",   14,  "center"),
    ("Py Net Transport",    13,  "center"),
    ("Py Fuel",             10,  "center"),
    ("Py Residential",      12,  "center"),
    ("Py DAS",               9,  "center"),
    ("Py Add'l Handling",   14,  "center"),
    ("Py Large Pkg",        10,  "center"),
    ("Py Remote Area",      11,  "center"),
    ("Py Declared Val",     12,  "center"),
    ("Py Addr Corr",        11,  "center"),
    ("Py Total",            11,  "center"),
    # Comparison
    ("Diff $",               8,  "center"),
    ("Match",                8,  "center"),
]

SERVICE_LABELS = {
    "ground":    "UPS Ground",
    "3day":      "UPS 3 Day Select",
    "2day":      "UPS 2nd Day Air",
    "2day_am":   "UPS 2nd Day Air A.M.",
    "nda_saver": "UPS NDA Saver",
    "nda":       "UPS Next Day Air",
}

FEDEX_SERVICE_LABELS = {
    "ground": "FedEx Ground",
    "home_delivery": "FedEx Home Delivery",
    "express_saver": "FedEx Express Saver",
    "2day": "FedEx 2Day",
    "standard_overnight": "FedEx Standard Overnight",
    "priority_overnight": "FedEx Priority Overnight",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _dims_str(d: dict | None) -> str:
    if not d:
        return "—"
    return f"{d['length']}×{d['width']}×{d['height']}"

def _flags_str(inp: dict) -> str:
    flags = []
    if inp.get("nonStandardPackaging"):
        flags.append("NonStd")
    if inp.get("addressCorrection"):
        flags.append("AddrCorr")
    return ", ".join(flags) if flags else "—"

def _money(v: float | None) -> str:
    if v is None or v == 0:
        return "—"
    return f"${v:,.2f}"

def _pct(v: float) -> str:
    return f"{v * 100:.1f}%"

# ---------------------------------------------------------------------------
# Build rows
# ---------------------------------------------------------------------------

def build_rows(cases: list[dict], tolerance: float = 0.01, carrier: str = "ups") -> list[dict]:
    estimate = estimate_fedex if carrier == "fedex" else estimate_ups
    rows = []
    for i, case in enumerate(cases, 1):
        inp = case["input"]
        tool_result = case.get("tool_result", {})
        markup_pct = float(case.get("markupPct", 0))
        expect_error = bool(case.get("expectError", False))

        result = estimate(inp)

        dims_str   = _dims_str(inp.get("dimensionsIn"))
        flags_str  = _flags_str(inp)
        declared   = inp.get("declaredValueDollars", 0) or 0
        svc_key = inp["service"]
        service = (
            FEDEX_SERVICE_LABELS.get(svc_key, svc_key)
            if carrier == "fedex"
            else SERVICE_LABELS.get(svc_key, svc_key)
        )
        program    = "Small Business" if inp.get("rateType") == "smallBusiness" else ("Daily" if carrier == "ups" else "List")
        delivery   = "Residential" if inp.get("residential") else "Commercial"

        tool_total = tool_result.get("totalEstimatedCharge")

        if expect_error:
            if not result["ok"]:
                rows.append({
                    "tc": i, "origin": inp.get("zoneChartPrefix"), "dest": inp["destinationZip"],
                    "service": service, "program": program, "delivery": delivery,
                    "weight": inp["weightLbs"], "dims": dims_str, "flags": flags_str,
                    "declared": declared or None, "markup": markup_pct or None,
                    "tool_zone": "N/A", "tool_billwt": "N/A", "tool_total": "N/A",
                    "py_zone": "N/A", "py_billwt": "N/A",
                    "py_pub": None, "py_net": None, "py_fuel": None,
                    "py_res": None, "py_das": None, "py_ah": None,
                    "py_lp": None, "py_remote": None, "py_dv": None, "py_ac": None,
                    "py_total": None, "diff": None, "match": "PASS (expected error)",
                    "status": "pass",
                })
            else:
                rows.append({
                    "tc": i, "origin": inp.get("zoneChartPrefix"), "dest": inp["destinationZip"],
                    "service": service, "program": program, "delivery": delivery,
                    "weight": inp["weightLbs"], "dims": dims_str, "flags": flags_str,
                    "declared": declared or None, "markup": markup_pct or None,
                    "tool_zone": "N/A", "tool_billwt": "N/A", "tool_total": "N/A",
                    "py_zone": result["breakdown"]["zone"], "py_billwt": result["breakdown"]["billableWeightLbs"],
                    "py_pub": None, "py_net": None, "py_fuel": None,
                    "py_res": None, "py_das": None, "py_ah": None,
                    "py_lp": None, "py_remote": None, "py_dv": None, "py_ac": None,
                    "py_total": result["breakdown"]["totalEstimatedCharge"],
                    "diff": None, "match": "FAIL (expected error)", "status": "fail",
                })
            continue

        if not result["ok"]:
            rows.append({
                "tc": i, "origin": inp.get("zoneChartPrefix"), "dest": inp["destinationZip"],
                "service": service, "program": program, "delivery": delivery,
                "weight": inp["weightLbs"], "dims": dims_str, "flags": flags_str,
                "declared": declared or None, "markup": markup_pct or None,
                "tool_zone": "?", "tool_billwt": "?",
                "tool_total": tool_total,
                "py_zone": "ERR", "py_billwt": "ERR",
                "py_pub": None, "py_net": None, "py_fuel": None,
                "py_res": None, "py_das": None, "py_ah": None,
                "py_lp": None, "py_remote": None, "py_dv": None, "py_ac": None,
                "py_total": None, "diff": None, "match": f"ERROR: {result['error']}", "status": "error",
            })
            continue

        bd = result["breakdown"]
        py_total = bd["totalEstimatedCharge"]
        diff = abs(py_total - tool_total) if tool_total is not None else None
        ok = diff is not None and diff <= tolerance

        rows.append({
            "tc": i,
            "origin": inp.get("zoneChartPrefix", inp.get("destinationZip", "?")[:3]),
            "dest": inp["destinationZip"],
            "service": service, "program": program, "delivery": delivery,
            "weight": inp["weightLbs"], "dims": dims_str, "flags": flags_str,
            "declared": declared or None, "markup": markup_pct or None,
            "tool_zone": "?",
            "tool_billwt": "?",
            "tool_total": tool_total,
            "py_zone": bd["zone"],
            "py_billwt": f"{bd['billableWeightLbs']} lb ({bd['billableWeightSource'][:3].upper()})",
            "py_pub":    bd["publishedRate"],
            "py_net":    bd["netTransportationCharge"],
            "py_fuel":   bd.get("fuelSurcharge") or None,
            "py_res":    (bd.get("residentialSurcharge") or 0) + (bd.get("homeDeliverySurcharge") or 0) or None,
            "py_das":    bd.get("dasSurcharge") or None,
            "py_ah":     bd.get("additionalHandlingSurcharge") or None,
            "py_lp":     bd.get("largePackageSurcharge") or bd.get("oversizeSurcharge") or None,
            "py_remote": bd.get("remoteAreaSurcharge") or None,
            "py_dv":     bd.get("declaredValueCharge") or None,
            "py_ac":     bd.get("addressCorrectionCharge") or None,
            "py_total":  py_total,
            "diff":      diff,
            "match":     "PASS" if ok else "FAIL",
            "status":    "pass" if ok else "fail",
        })
    return rows

# ---------------------------------------------------------------------------
# Write Excel
# ---------------------------------------------------------------------------

def write_excel(rows: list[dict], out_path: Path, title: str = "UPS Rate Estimator — Python Cross-Validation") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing Audit"
    ws.freeze_panes = "A3"  # freeze title + header rows

    # ---- Title row ----
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", bold=True, size=13, color=HEADER_TXT)
    title_cell.fill = _fill(NAVY)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 22

    # ---- Group header bands ----
    # inputs: cols 1-11, tool: 12-14, python: 15-27, result: 28-29
    group_bands = [
        (1,  11, "INPUTS",          NAVY_LIGHT),
        (12, 14, "TOOL OUTCOME",    "2E6DA4"),
        (15, 27, "PYTHON OUTCOME",  "3A7D44"),
        (28, 29, "COMPARISON",      "7B3F8C"),
    ]
    for start, end, label, color in group_bands:
        ws.merge_cells(f"{get_column_letter(start)}2:{get_column_letter(end)}2")
        c = ws.cell(row=2, column=start, value=label)
        c.font = Font(name="Calibri", bold=True, size=9, color=HEADER_TXT)
        c.fill = _fill(color)
        c.alignment = _center()
    ws.row_dimensions[2].height = 16

    # ---- Column headers (row 3) ----
    ws.row_dimensions[3].height = 30
    for col_i, (hdr, width, align) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=col_i, value=hdr)
        c.font = Font(name="Calibri", bold=True, size=9, color=HEADER_TXT)
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(col_i)].width = width

    # ---- Data rows ----
    for row_i, row in enumerate(rows, 4):
        status = row["status"]
        bg = GREEN_PASS if status == "pass" else RED_FAIL if status == "fail" else AMBER_ERR
        alt = GRAY_ROW if (row_i % 2 == 0) else WHITE

        def cell(col: int, value, money: bool = False, pct: bool = False, align: str = None) -> None:
            c = ws.cell(row=row_i, column=col)
            if value is None or value == 0 and money:
                c.value = None
            elif money and isinstance(value, (int, float)):
                c.value = value
                c.number_format = '"$"#,##0.00'
            elif pct and isinstance(value, (int, float)):
                c.value = f"{value:.0f}%"
            else:
                c.value = value
            c.font = _font(size=9)
            col_align = COLS[col - 1][2] if align is None else align
            c.alignment = Alignment(horizontal=col_align, vertical="center")
            c.border = _border()
            # Row background: result columns get status color, others get alternating
            if col in (28, 29):
                c.fill = _fill(bg)
                if col == 29:
                    txt = GREEN_TXT if status == "pass" else RED_TXT if status == "fail" else AMBER_TXT
                    c.font = Font(name="Calibri", bold=True, size=9, color=txt)
            else:
                c.fill = _fill(alt)

        ws.row_dimensions[row_i].height = 15

        cell(1,  row["tc"])
        cell(2,  row["origin"])
        cell(3,  row["dest"])
        cell(4,  row["service"], align="left")
        cell(5,  row["program"])
        cell(6,  row["delivery"])
        cell(7,  row["weight"])
        cell(8,  row["dims"])
        cell(9,  row["flags"], align="left")
        cell(10, row["declared"], money=True)
        cell(11, row["markup"], pct=True)
        cell(12, row["tool_zone"])
        cell(13, row["tool_billwt"])
        cell(14, row["tool_total"], money=True)
        cell(15, row["py_zone"])
        cell(16, row["py_billwt"])
        cell(17, row["py_pub"],    money=True)
        cell(18, row["py_net"],    money=True)
        cell(19, row["py_fuel"],   money=True)
        cell(20, row["py_res"],    money=True)
        cell(21, row["py_das"],    money=True)
        cell(22, row["py_ah"],     money=True)
        cell(23, row["py_lp"],     money=True)
        cell(24, row["py_remote"], money=True)
        cell(25, row["py_dv"],     money=True)
        cell(26, row["py_ac"],     money=True)
        cell(27, row["py_total"],  money=True)
        cell(28, row["diff"],      money=True)
        cell(29, row["match"])

    # ---- Summary row ----
    summary_row = len(rows) + 4
    ws.row_dimensions[summary_row].height = 16
    passed = sum(1 for r in rows if r["status"] == "pass")
    failed = sum(1 for r in rows if r["status"] == "fail")
    errors = sum(1 for r in rows if r["status"] == "error")

    ws.merge_cells(f"A{summary_row}:M{summary_row}")
    sc = ws.cell(row=summary_row, column=1,
                 value=f"Summary: {passed} PASS  |  {failed} FAIL  |  {errors} ERROR  |  {len(rows)} total cases")
    sc.font = Font(name="Calibri", bold=True, size=10, color=HEADER_TXT)
    sc.fill = _fill(NAVY if failed == 0 else "8B0000")
    sc.alignment = _center()
    for col in range(14, len(COLS) + 1):
        ws.cell(row=summary_row, column=col).fill = _fill(NAVY if failed == 0 else "8B0000")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}  ({passed}/{len(rows)} passed)")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--carrier", choices=["ups", "fedex"], default="ups")
    parser.add_argument("--tests", default=str(Path(__file__).parent / "pricing_test_cases.json"))
    parser.add_argument("--out",   default=str(Path(__file__).parent.parent / "outputs" / "pricing_audit.xlsx"))
    parser.add_argument("--tolerance", type=float, default=0.01)
    args = parser.parse_args()

    with open(args.tests) as f:
        cases = json.load(f)

    title = (
        "FedEx Rate Estimator — Python Cross-Validation"
        if args.carrier == "fedex"
        else "UPS Rate Estimator — Python Cross-Validation"
    )
    rows = build_rows(cases, args.tolerance, carrier=args.carrier)
    write_excel(rows, Path(args.out), title=title)
